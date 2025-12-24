import streamlit as st
import pandas as pd
import os
import re
import zipfile
import tempfile
import io
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --- Helper Functions ---
def excel_col_to_index(col_str):
    num = 0
    for c in col_str:
        if c in "0123456789":
            continue
        num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num - 1

def convert_implied_decimal(val):
    try:
        val_str = str(val).strip()
        return float(val_str) / 100.0
    except:
        return val

def extract_seq_num(val):
    text = str(val)
    match = re.search(r'seq_num:(\d+)', text)
    if match:
        return match.group(1)
    return str(val).strip()

def parse_dates_from_filename(filename: str):
    base = os.path.basename(filename)
    d_match = re.search(r'[-_]?D(?P<d>\d{6})', base, flags=re.IGNORECASE)
    d_date = d_match.group('d') if d_match else None

    jv_match = re.search(r'JV(?P<jv>\d{8})', base, flags=re.IGNORECASE)
    jv_date = None
    if jv_match:
        jv_full = jv_match.group('jv')
        jv_date = jv_full[2:4] + jv_full[4:6] + jv_full[6:8] 
    return d_date, jv_date

def pick_latest_files_by_duplicate_d_date(folder_path: str, files_list: list):
    """
    files_list: รายชื่อไฟล์ที่เป็น Source (ไม่ใช่ TLF)
    """
    chosen = {} 

    for fn in files_list:
        file_path = os.path.join(folder_path, fn)
        
        # รองรับนามสกุลไฟล์
        valid_exts = ('.csv', '.trf', '.txt', '.xls', '.xlsx')
        if not fn.lower().endswith(valid_exts):
            continue

        d_date, jv_date = parse_dates_from_filename(fn)
        jv_int = int(jv_date) if jv_date and jv_date.isdigit() else -1

        if d_date is None:
            key = f"__NO_D__::{fn}"
            chosen[key] = {"file": file_path, "d_date": None, "jv_date": jv_date}
            continue

        if d_date not in chosen:
            chosen[d_date] = {"file": file_path, "d_date": d_date, "jv_date": jv_date, "_jv_int": jv_int}
        else:
            if jv_int > chosen[d_date].get("_jv_int", -1):
                chosen[d_date] = {"file": file_path, "d_date": d_date, "jv_date": jv_date, "_jv_int": jv_int}

    results = []
    for _, v in chosen.items():
        v.pop("_jv_int", None)
        results.append(v)

    results.sort(key=lambda x: os.path.basename(x["file"]).lower())
    return results

def strip_d_suffix_for_tlf_sheet(name_no_ext: str):
    return re.sub(r'[-_]?D\d{6}.*$', '', name_no_ext, flags=re.IGNORECASE).strip()

def make_unique_sheet_name(book, desired_name: str):
    base = (desired_name or "Sheet")[:31]
    name = base
    i = 2
    while name in book.sheetnames:
        suffix = f"_{i}"
        name = (base[:31 - len(suffix)] + suffix)[:31]
        i += 1
    return name

# --- Configuration Constants ---
tlf_reserved_rows = 2
gl_reserved_rows = 10
gap_rows = 3
exclude_tlf_columns = ['from_acct', 'to_acct', 'auth_branch_from']

gl_columns_letters = ['J', 'K', 'L', 'M', 'N', 'P', 'AM', 'AN', 'AZ']
gl_new_headers = ['RC', 'OC', 'CH', 'Product Code', 'Account Code', 'Tax', 'DR', 'CR', 'Seq']

tlf_columns_letters = [
    'F', 'G', 'I', 'J', 'K', 'M', 'O', 'V',
    'AF', 'AS', 'AT', 'AU', 'AV', 'AX', 'AZ', 'CU', 'DP', 'BH'
]

gl_indices = [excel_col_to_index(c) for c in gl_columns_letters]
tlf_indices = [excel_col_to_index(c) for c in tlf_columns_letters]

def get_col_pos_in_tlf(target_letter):
    sorted_letters = sorted(tlf_columns_letters, key=lambda x: excel_col_to_index(x))
    try:
        return sorted_letters.index(target_letter)
    except:
        return -1

pos_AZ = get_col_pos_in_tlf('AZ')
pos_CU = get_col_pos_in_tlf('CU')

# Styles
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
align_center = Alignment(horizontal='center', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')
header_font = Font(bold=True)
title_font = Font(bold=True, size=14, color="000000")
search_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


# --- Main Logic Function ---
def process_data_in_memory(tlf_path, source_files_list, temp_folder):
    output = io.BytesIO()
    
    # [FIX] ใช้ with block เพื่อเปิดและปิดไฟล์ TLF อัตโนมัติ ป้องกัน WinError 32
    try:
        with pd.ExcelFile(tlf_path) as tlf_book:
            
            # 2. Filter Source Files
            files_to_process = pick_latest_files_by_duplicate_d_date(temp_folder, source_files_list)
            if not files_to_process:
                return None, "ไม่พบไฟล์ข้อมูล (GL/TRF) ที่ถูกต้องใน ZIP"

            # 3. Create Excel Writer
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                for item in files_to_process:
                    file_path = item["file"]
                    filename = os.path.basename(file_path)
                    chosen_d_date = item["d_date"]

                    # Sheet Name
                    desired_sheet_name = chosen_d_date if chosen_d_date else os.path.splitext(filename)[0]
                    
                    # Lookup Logic
                    clean_name = re.sub(r'GL', '', filename, flags=re.IGNORECASE)
                    clean_name = os.path.splitext(clean_name)[0].strip()
                    fallback_lookup_name = strip_d_suffix_for_tlf_sheet(clean_name)

                    tlf_lookup_candidates = []
                    if chosen_d_date:
                        tlf_lookup_candidates.append(chosen_d_date)
                        tlf_lookup_candidates.append("D" + chosen_d_date)
                    tlf_lookup_candidates.append(fallback_lookup_name)

                    tlf_sheet_to_use = None
                    for cand in tlf_lookup_candidates:
                        if cand and cand in tlf_book.sheet_names:
                            tlf_sheet_to_use = cand
                            break

                    try:
                        # --- Load TLF Data ---
                        tlf_df = pd.DataFrame()
                        if tlf_sheet_to_use:
                            tlf_df = pd.read_excel(tlf_book, sheet_name=tlf_sheet_to_use, usecols=tlf_indices, dtype=str)
                            for col in tlf_df.columns:
                                tlf_df[col] = tlf_df[col].astype(str).str.strip()

                            if pos_AZ != -1 and pos_AZ < len(tlf_df.columns):
                                tlf_df.iloc[:, pos_AZ] = tlf_df.iloc[:, pos_AZ].apply(convert_implied_decimal)
                            if pos_CU != -1 and pos_CU < len(tlf_df.columns):
                                tlf_df.iloc[:, pos_CU] = tlf_df.iloc[:, pos_CU].apply(convert_implied_decimal)

                            if not tlf_df.empty and len(tlf_df.columns) > 8:
                                search_col = tlf_df.iloc[:, 8].astype(str).str.strip()
                                tlf_df['_SearchKey'] = search_col + '|' + (tlf_df.groupby(search_col).cumcount() + 1).astype(str)

                        # --- Load Source Data (GL/TRF) ---
                        # [FIX] ใช้ with block สำหรับ Source file ที่เป็น Excel ด้วย
                        if filename.lower().endswith(('.xls', '.xlsx')):
                             with pd.ExcelFile(file_path) as source_book:
                                gl_df = pd.read_excel(source_book, header=None, usecols=gl_indices, dtype=str)
                        else:
                            try:
                                gl_df = pd.read_csv(file_path, header=None, usecols=gl_indices, encoding='utf-8', dtype=str, engine='python')
                            except:
                                gl_df = pd.read_csv(file_path, header=None, usecols=gl_indices, encoding='cp874', dtype=str, engine='python')

                        if len(gl_df.columns) == len(gl_new_headers):
                            gl_df.columns = gl_new_headers

                        gl_df['RC'] = gl_df['RC'].astype(str).str.strip()
                        gl_df['CH'] = gl_df['CH'].astype(str).str.strip()
                        gl_df['DR'] = pd.to_numeric(gl_df['DR'], errors='coerce').fillna(0)
                        gl_df['CR'] = pd.to_numeric(gl_df['CR'], errors='coerce').fillna(0)

                        if 'Seq' in gl_df.columns:
                            gl_df['Seq'] = gl_df['Seq'].apply(extract_seq_num).astype(str).str.strip()

                        cols_to_sort = ['CH', 'RC', 'OC', 'Product Code']
                        valid_sort_cols = [c for c in cols_to_sort if c in gl_df.columns]
                        if valid_sort_cols:
                            gl_df = gl_df.sort_values(by=valid_sort_cols, ascending=[True]*len(valid_sort_cols))

                        if not gl_df.empty:
                            search_col_gl = gl_df['Seq'].astype(str)
                            gl_df['_SearchKey'] = search_col_gl + '|' + (gl_df.groupby(search_col_gl).cumcount() + 1).astype(str)

                        # --- Write to Excel (Layout Logic) ---
                        target_sheet_name = make_unique_sheet_name(writer.book, desired_sheet_name)
                        worksheet = writer.book.create_sheet(target_sheet_name)
                        writer.sheets[target_sheet_name] = worksheet
                        ws = writer.sheets[target_sheet_name]

                        search_ui_start_row = 1
                        tlf_ui_height = 2 + tlf_reserved_rows
                        gl_ui_height = 2 + gl_reserved_rows
                        raw_data_start_row = search_ui_start_row + tlf_ui_height + gap_rows + gl_ui_height + 5

                        current_raw_row = raw_data_start_row
                        tlf_data_start = tlf_data_end = None
                        tlf_key_col_letter = 'A'
                        gl_data_start = gl_data_end = None
                        gl_key_col_letter = 'A'

                        # Raw TLF
                        if not tlf_df.empty:
                            ws.cell(row=current_raw_row - 1, column=1, value="--- Raw TLF Data ---").font = Font(bold=True, italic=True)
                            tlf_df.to_excel(writer, sheet_name=target_sheet_name, startrow=current_raw_row - 1, index=False)
                            tlf_data_start = current_raw_row + 1
                            tlf_data_end = current_raw_row + len(tlf_df)
                            tlf_key_col_letter = get_column_letter(len(tlf_df.columns))
                            
                            for row in range(current_raw_row, tlf_data_end + 1):
                                for col in range(1, len(tlf_df.columns)):
                                    cell = ws.cell(row=row, column=col)
                                    cell.border = thin_border
                                    if row == current_raw_row:
                                        cell.alignment = align_center
                                        cell.font = header_font
                                    else:
                                        cell.alignment = align_right if isinstance(cell.value, (int, float)) else align_center
                                        if col == 9: cell.number_format = '@'
                            current_raw_row += len(tlf_df) + 4

                        # Raw GL
                        if not gl_df.empty:
                            ws.cell(row=current_raw_row - 1, column=1, value="--- Raw ATMI Data ---").font = Font(bold=True, italic=True)
                            gl_df.to_excel(writer, sheet_name=target_sheet_name, startrow=current_raw_row - 1, index=False)
                            gl_data_start = current_raw_row + 1
                            gl_data_end = current_raw_row + len(gl_df)
                            gl_key_col_letter = get_column_letter(len(gl_df.columns))

                            for row in range(current_raw_row, gl_data_end + 1):
                                for col in range(1, len(gl_df.columns) + 1):
                                    cell = ws.cell(row=row, column=col)
                                    cell.border = thin_border
                                    if row == current_raw_row:
                                        cell.alignment = align_center
                                        cell.font = header_font
                                    else:
                                        if col in [7, 8]:
                                            cell.alignment = align_right
                                            cell.number_format = '#,##0.00'
                                        else:
                                            cell.alignment = align_center
                                        if col == 9: cell.number_format = '@'

                        # --- Search UI ---
                        ws[f'A{search_ui_start_row}'] = "🔍 ค้นหาข้อมูล SEQ"
                        ws[f'A{search_ui_start_row}'].font = Font(bold=True, size=12)
                        ws[f'A{search_ui_start_row}'].alignment = Alignment(horizontal='right')

                        input_cell_ref = f'$B${search_ui_start_row}'
                        input_cell = ws[f'B{search_ui_start_row}']
                        input_cell.fill = search_fill
                        input_cell.border = thin_border
                        input_cell.alignment = align_center
                        input_cell.number_format = '@'

                        report_row = search_ui_start_row + 2

                        # TLF Report Section
                        if not tlf_df.empty:
                            ws[f'A{report_row}'] = "TLF"
                            ws[f'A{report_row}'].font = title_font
                            
                            display_cols = [c for c in tlf_df.columns if c != '_SearchKey' and c not in exclude_tlf_columns]
                            if 'amt_1_full' in display_cols and 'resp_byte' in display_cols:
                                idx1, idx2 = display_cols.index('amt_1_full'), display_cols.index('resp_byte')
                                display_cols[idx1], display_cols[idx2] = display_cols[idx2], display_cols[idx1]

                            current_col_idx = 1
                            tlf_key_range_str = f"${tlf_key_col_letter}${tlf_data_start}:${tlf_key_col_letter}${tlf_data_end}"

                            for col_name in display_cols:
                                cell = ws.cell(row=report_row + 1, column=current_col_idx)
                                cell.value = col_name
                                cell.font = Font(bold=True)
                                cell.border = thin_border
                                cell.alignment = align_center
                                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                                current_col_idx += 1

                            data_start_row = report_row + 2
                            for r_offset in range(tlf_reserved_rows):
                                current_formula_row = data_start_row + r_offset
                                k_value = r_offset + 1
                                match_logic = f'MATCH({input_cell_ref}&"|"&{k_value}, {tlf_key_range_str}, 0)'

                                for i, col_name in enumerate(display_cols, 1):
                                    original_col_idx = tlf_df.columns.get_loc(col_name)
                                    col_letter = get_column_letter(original_col_idx + 1)
                                    data_col_range = f"${col_letter}${tlf_data_start}:${col_letter}${tlf_data_end}"
                                    formula = f'=IFERROR(INDEX({data_col_range}, {match_logic}), "")'
                                    
                                    cell = ws.cell(row=current_formula_row, column=i)
                                    cell.value = formula
                                    cell.border = thin_border
                                    cell.alignment = align_center

                            report_row = data_start_row + tlf_reserved_rows

                        report_row += gap_rows

                        # GL Report Section
                        if not gl_df.empty:
                            ws[f'A{report_row}'] = "ATMI"
                            ws[f'A{report_row}'].font = title_font
                            
                            current_col_idx = 1
                            for col_name in gl_df.columns:
                                if col_name == '_SearchKey': continue
                                cell = ws.cell(row=report_row + 1, column=current_col_idx)
                                cell.value = col_name
                                cell.font = Font(bold=True)
                                cell.border = thin_border
                                cell.alignment = align_center
                                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                                current_col_idx += 1
                            
                            data_start_row = report_row + 2
                            gl_key_range_str = f"${gl_key_col_letter}${gl_data_start}:${gl_key_col_letter}${gl_data_end}"

                            for r_offset in range(gl_reserved_rows):
                                current_formula_row = data_start_row + r_offset
                                k_value = r_offset + 1
                                match_logic = f'MATCH({input_cell_ref}&"|"&{k_value}, {gl_key_range_str}, 0)'
                                
                                for col_idx in range(1, len(gl_new_headers) + 1):
                                    col_letter = get_column_letter(col_idx)
                                    data_col_range = f"${col_letter}${gl_data_start}:${col_letter}${gl_data_end}"
                                    formula = f'=IFERROR(INDEX({data_col_range}, {match_logic}), "")'
                                    
                                    cell = ws.cell(row=current_formula_row, column=col_idx)
                                    cell.value = formula
                                    cell.border = thin_border
                                    if col_idx in [7, 8]:
                                        cell.number_format = '#,##0.00'
                                        cell.alignment = align_right
                                    else:
                                        cell.alignment = align_center

                        # Auto Width
                        col_widths = {}
                        def update_max_width(df, start_col_idx=1):
                            for i, col_name in enumerate(df.columns):
                                current_idx = start_col_idx + i
                                max_len = len(str(col_name))
                                if not df.empty:
                                    try:
                                        data_len = df[col_name].astype(str).map(len).max()
                                        if pd.notna(data_len): max_len = max(max_len, data_len)
                                    except: pass
                                existing = col_widths.get(current_idx, 0)
                                col_widths[current_idx] = max(existing, max_len + 3)

                        if not tlf_df.empty: update_max_width(tlf_df)
                        if not gl_df.empty: update_max_width(gl_df)

                        for col_idx, width in col_widths.items():
                            col_letter = get_column_letter(col_idx)
                            writer.sheets[target_sheet_name].column_dimensions[col_letter].width = max(12, min(width, 60))
                        
                        writer.sheets[target_sheet_name].column_dimensions['A'].width = max(col_widths.get(1, 20), 30)

                    except Exception as e:
                        print(f"Error processing file {filename}: {e}")

                if 'Sheet' in writer.book.sheetnames and len(writer.book.sheetnames) > 1:
                    del writer.book['Sheet']

    except Exception as e:
        return None, f"Error อ่านไฟล์ TLF: {e}"
    
    output.seek(0)
    return output, None

# --- Streamlit UI ---
st.title("📂 Automated GL & TLF Matching Report")
st.write("อัปโหลดไฟล์ ZIP ที่ประกอบด้วยไฟล์ข้อมูล (`.trf`/`.csv`) และไฟล์ `TLF`")

uploaded_zip = st.file_uploader("Choose a ZIP file", type="zip")

if uploaded_zip:
    if st.button("🚀 Process Files"):
        with st.spinner("Extracting & Processing..."):
            # Create temp directory
            with tempfile.TemporaryDirectory() as temp_dir:
                try:
                    # Extract ZIP
                    with zipfile.ZipFile(uploaded_zip, 'r') as zip_ref:
                        zip_ref.extractall(temp_dir)
                    
                    # Identify Files
                    tlf_path = None
                    source_files = []
                    
                    for root, dirs, files in os.walk(temp_dir):
                        for file in files:
                            if file.startswith('.') or '__MACOSX' in root:
                                continue
                            
                            full_path = os.path.join(root, file)
                            
                            if "TLF" in file:
                                tlf_path = full_path
                            else:
                                source_files.append(file)
                    
                    # Process
                    if not tlf_path:
                        st.error("❌ ไม่พบไฟล์ TLF ใน ZIP (ต้องมีคำว่า 'TLF' ในชื่อไฟล์)")
                    elif not source_files:
                        st.error("❌ ไม่พบไฟล์ข้อมูล Source (CSV/TRF) ใน ZIP")
                    else:
                        st.info(f"📍 Found TLF: {os.path.basename(tlf_path)}")
                        st.info(f"📍 Found Source Files: {len(source_files)} files")
                        
                        excel_file, error_msg = process_data_in_memory(tlf_path, source_files, temp_dir)
                        
                        if error_msg:
                            st.error(error_msg)
                        else:
                            st.success("✅ Processing Complete!")
                            st.download_button(
                                label="📥 Download Final Excel",
                                data=excel_file,
                                file_name="Combined_GL_Final_V4_Reorder.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                            
                except Exception as e:
                    st.error(f"Error during processing: {e}")
